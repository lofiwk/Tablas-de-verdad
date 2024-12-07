import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Obtener el directorio donde se encuentra el script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Definir entradas (S1, S2, S3, S4, S5)
entradas = ["S1", "S2", "S3", "S4", "S5"]

# Generar todas las combinaciones de entradas (32 combinaciones para 5 bits)
combinaciones = np.array(np.meshgrid([0, 1], [0, 1], [0, 1], [0, 1], [0, 1])).T.reshape(-1, 5)

# Crear DataFrame con las combinaciones
tabla_verdad = pd.DataFrame(combinaciones, columns=entradas)

# Función para calcular las salidas universales según el enunciado "a"
def calcular_salidas(filas):
    S1, S2, S3, S4, S5 = filas
    if S1 == S2 == S3 == S4 == S5:  # Todos iguales
        return [1, 1, 1, 1]  # Todas encendidas
    else:
        return [0, 0, 0, 0]  # Todas apagadas

# Aplicar las reglas para calcular las salidas universales
tabla_verdad[["B1", "B2", "B3", "B4"]] = tabla_verdad.apply(calcular_salidas, axis=1, result_type="expand")

# Añadir explicaciones textuales de los casos
tabla_verdad["Identificador_Casos"] = (
    "Caso A: Sin cambios. | "
    "Caso B: Invertir entradas (S1-S5). | "
    "Caso C: Invertir salidas (B1-B4). | "
    "Caso D: Invertir entradas y salidas."
)

# Crear archivo Excel con formato
wb = Workbook()
ws = wb.active
ws.title = "Tabla de Verdad Universal"

# Escribir encabezados con estilo
encabezados = list(tabla_verdad.columns)
for col, encabezado in enumerate(encabezados, start=1):
    cell = ws.cell(row=1, column=col, value=encabezado)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

# Escribir datos
for row_idx, row in enumerate(dataframe_to_rows(tabla_verdad, index=False, header=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

# Ajustar ancho de columnas
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[col[0].column_letter].width = max_length + 4

# Ajustar altura de filas
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 20

# Guardar archivo en el mismo directorio que el script
output_path = os.path.join(script_dir, "tabla_verdad_a_con_casos.xlsx")
wb.save(output_path)
print(f"Archivo Excel creado: {output_path}")
